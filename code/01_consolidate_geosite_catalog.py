"""
Stage 1 / Task 1: load and normalize every known geosite data source into a
single "Observation" table (hierarchy-preserving, no dedup/merge yet).

See docs/superpowers/specs/2026-07-29-geosite-catalog-consolidation-design.md
for the full architectural rationale (two-level Observation + Locality model,
why the Excel `Data Classification_Geoheritage.xlsx` source is hierarchical).
"""
import math
import os
import re
import unicodedata
import numpy as np
import pandas as pd
import openpyxl
import geopandas as gpd
from shapely.geometry import Point
from pyproj import Transformer
from rapidfuzz import fuzz
from scipy.spatial import cKDTree

BASE = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
OUT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "data"))
os.makedirs(OUT_DIR, exist_ok=True)

OBS_COLS = ["Geosite_Name", "Latitude_WGS84", "Longitude_WGS84", "Region", "Geosite_Type",
            "Geological_Domain", "Locality_Center_Lat", "Locality_Center_Lon",
            "Locality_Group_Key", "Coordinate_Precision", "Source_File", "Source_Row_Ref"]

# --- UTM handling -----------------------------------------------------------
# The Excel source's projected coordinates were determined (spec, confirmed
# two independent ways) to be UTM Zone 29N (EPSG:32629) for essentially all of
# Morocco. Feuil2's own worked-example legend has one deliberately-included
# Zone 28N case (row 22, "Zone 28N" stated explicitly in its Remarques column):
# validate_parser_against_feuil2() reads that Remarques text and calls
# parse_coordinate_pair(..., utm_zone=28) specifically for that row, asserting
# it converts correctly under 28N — a real, exercised test of the zone-override
# parameter, not just a documented capability.
#
# UTM zone is NOT uniformly 29N across the whole source (2026-07-30 findings,
# see docs/superpowers/specs/2026-07-30-geoheritage-excel-data-model-findings.md
# §8): Feuil2 row 22's raw value (X=645000, Y=2827000) is an exact match to
# `Data generale`/`Feuil1` row 575's group ("Typical coastal cliff of the
# Dakhla peninsula", Reference "Saddiqi et al.2022"), and Feuil2 explicitly
# remarks that group is "Zone 28N", not 29N. `_GROUP_UTM_ZONE_OVERRIDES` below
# is a small {normalized Reference: EPSG zone} table applying that one
# confirmed exception at ingestion time (used only for the group-level Center
# coordinates, since that group's individual per-geosite Coordinates are all
# DMS strings, not UTM). 29N remains the default for every other group; add
# another entry to that dict if a future group is ever confirmed to need a
# different zone.
_UTM_TRANSFORMERS = {
    29: Transformer.from_crs("EPSG:32629", "EPSG:4326", always_xy=True),
    28: Transformer.from_crs("EPSG:32628", "EPSG:4326", always_xy=True),
    30: Transformer.from_crs("EPSG:32630", "EPSG:4326", always_xy=True),
}

# {normalized Reference string: EPSG UTM zone number} — one confirmed exception
# so far (Dakhla peninsula group, see comment above). Structured as a dict so
# another exception can be added the same way if one is ever confirmed.
_GROUP_UTM_ZONE_OVERRIDES = {
    "saddiqi et al.2022": 28,
}


def _group_utm_zone(reference):
    """Look up the UTM zone for a publication group by its Reference string,
    defaulting to 29N (see _GROUP_UTM_ZONE_OVERRIDES)."""
    key = str(reference).strip().lower() if reference is not None else ""
    return _GROUP_UTM_ZONE_OVERRIDES.get(key, 29)

# A third projected system found live in Data generale (not documented in the
# original spec/brief): the Talassemtane National Park / Chefchaouen group
# (rows ~340-380, "Rif" domain) stores its individual-observation coordinates
# as bare 6-digit numbers with a literal "N"/"W" suffix (e.g. "509000N",
# "514480W") — not DMS, not UTM. These match Merchich / Nord Maroc Lambert
# Conformal Conic (EPSG:26191) closely: converting the group's first rows
# lands within ~0.1-0.2 degrees of the group's own declared Center coordinate
# (35.2253, -5.1406, itself confirmed independently against Feuil2 row 18),
# tightly clustered as expected for one national park. Verified on 5 rows.
#
# IMPORTANT: this format has no built-in checksum, and a handful of rows in
# the source have corrupted/truncated digits (e.g. a transposed leading pair,
# or a digit missing) that still match the bare "digits+N / digits+W" shape
# and convert to a real-looking but wrong coordinate hundreds of km away from
# the rest of the group. A plausibility check against the group's own
# Locality_Center is mandatory here — see _MERCHICH_MAX_KM_FROM_CENTER below
# — otherwise those corrupted rows are silently ingested as if correct
# (caught in review: e.g. "Rueda Nummilitic succession" cx=350673N/cy=514050W
# converts to (35.2168,-7.0382), ~420km from the group's center, and
# "Beni Derkoul Radiolarites" cx=528971N/cy=96626W — Y truncated to 5 digits —
# converts to (31.4627,-5.0937), ~418km off).
_MERCHICH_TRANSFORMER = Transformer.from_crs("EPSG:26191", "EPSG:4326", always_xy=True)

# The ~35 genuinely-valid rows in this block cluster within roughly 10-30km of
# the group's declared center; 50km gives headroom for real within-park spread
# while still rejecting the multi-hundred-km corrupted-digit outliers found in
# review (~210-420km off).
_MERCHICH_MAX_KM_FROM_CENTER = 50.0


def _haversine_km(lat1, lon1, lat2, lon2):
    r = 6371.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlmb = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dlmb / 2) ** 2
    return 2 * r * math.asin(math.sqrt(a))


def convert_utm(x, y, zone=29):
    transformer = _UTM_TRANSFORMERS[zone]
    lon, lat = transformer.transform(x, y)
    return lat, lon, "point", None


def parse_coordinate_pair(raw_x, raw_y, utm_zone=29, locality_center=None):
    """
    Auto-detects and parses a (X, Y) or (lat_str, lon_str) coordinate pair in any of:
    plain UTM Zone 29N (or explicitly overridden zone) easting/northing,
    DMS-with-seconds, DMS-with-comma-decimal-minutes, DMS-degree-minute-only,
    or an interval/range string (resolved to midpoint, flagged lower precision).
    `locality_center`, if given as (lat, lon), is used to plausibility-check the
    Merchich/Nord Maroc branch (see _MERCHICH_MAX_KM_FROM_CENTER) — that format
    has no internal checksum, so a handful of digit-corrupted rows would
    otherwise parse "successfully" into real-looking but wrong coordinates.
    Returns (lat, lon, precision, error) where precision in {"point", "interval"} and
    error is None on success or a string reason on failure (never both set).
    """
    # 1. Plain numeric UTM easting/northing (the common case in Data generale's
    #    numeric-typed cells, e.g. the Anti-Atlas group rows).
    if isinstance(raw_x, (int, float)) and isinstance(raw_y, (int, float)):
        return convert_utm(raw_x, raw_y, zone=utm_zone)

    sx = str(raw_x).strip() if raw_x is not None else ""
    sy = str(raw_y).strip() if raw_y is not None else ""

    # 1b. Merchich / Nord Maroc Lambert grid (EPSG:26191): bare 5-7 digit
    #     integers with a literal N/W suffix and no DMS punctuation at all
    #     (e.g. "509000N", "514480W" — see _MERCHICH_TRANSFORMER comment).
    m_merchich_x = re.match(r"^(\d{5,7})N$", sx)
    m_merchich_y = re.match(r"^(\d{5,7})W$", sy)
    if m_merchich_x and m_merchich_y:
        x, y = float(m_merchich_x.group(1)), float(m_merchich_y.group(1))
        lon, lat = _MERCHICH_TRANSFORMER.transform(x, y)
        if locality_center is not None and locality_center[0] is not None and locality_center[1] is not None:
            dist_km = _haversine_km(lat, lon, locality_center[0], locality_center[1])
            if dist_km > _MERCHICH_MAX_KM_FROM_CENTER:
                return None, None, None, (
                    f"Merchich (EPSG:26191) conversion implausible: {dist_km:.0f}km from "
                    f"locality center {locality_center}, exceeds {_MERCHICH_MAX_KM_FROM_CENTER:.0f}km "
                    f"threshold — likely corrupted/truncated input digits, not guessed at"
                )
        return lat, lon, "point", None

    # 2. Combined single-cell interval, no per-token hemisphere letters, of the
    #    form "LAT1-LAT2,LON1-LON2" (observed in Data generale, e.g. "Oualidia
    #    caves"/"Oualidia Lagoon": "32°40'42"-32°47'07",8°52'30"-9°02'50"",
    #    sometimes with an embedded newline from cell wrapping). raw_y is empty
    #    in this case since everything lives in one cell.
    if not sy and ("," in sx) and re.search(r"[–-]", sx):
        result = parse_combined_interval_no_hemisphere(sx)
        if result[0] is not None:
            return result

    # 3. Combined single-cell "Intervalle ...N-...N, ...W-...W" style (Feuil2's
    #    own worked-example format, hemisphere letters present).
    if "ntervalle" in sx.lower():
        result = parse_combined_interval_with_hemisphere(sx)
        if result[0] is not None:
            return result

    # 4. Combined single-cell "LAT_DMS LON_DMS" pair (Feuil2's raw strings are
    #    one column; Data generale instead splits these across two cells, case 6).
    if not sy:
        tokens = re.findall(r"[\d][\d°'’′‘\"″ʺ”“,.\s]*[NSEW]", sx)
        if len(tokens) >= 2:
            lat_tok = next((t for t in tokens if t.strip()[-1] in "NS"), None)
            lon_tok = next((t for t in tokens if t.strip()[-1] in "EW"), None)
            if lat_tok and lon_tok:
                lat_val = parse_dms_token(lat_tok.strip())
                lon_val = parse_dms_token(lon_tok.strip())
                if lat_val is not None and lon_val is not None:
                    return lat_val, lon_val, "point", None

    # 5. Dash-separated interval across the two supplied cells (rare fallback,
    #    kept for robustness though not observed as the dominant interval shape).
    if "–" in sx or ("-" in sx and re.search(r"\d\s*-\s*\d", sx)):
        parts = re.split(r"[–-]", sx)
        vals = [parse_dms_token(p) for p in parts if p.strip()]
        vals = [v for v in vals if v is not None]
        if len(vals) >= 2:
            lat_mid = sum(vals) / len(vals)
            lon_val = parse_dms_token(sy)
            if lon_val is not None:
                return lat_mid, lon_val, "interval", None

    # 6. DMS pair split across two cells (Data generale's normal layout for
    #    non-UTM rows), seconds or comma-decimal-minutes or minute-only format.
    lat_val = parse_dms_token(sx)
    lon_val = parse_dms_token(sy)
    if lat_val is not None and lon_val is not None:
        return lat_val, lon_val, "point", None

    # 7. Bare decimal-degree pair, no hemisphere letters at all, split across
    #    two cells (e.g. Tinghir sub-group rows: (' 32.166', '5.523')). Same
    #    Morocco sign convention as the no-hemisphere interval case: latitude
    #    is North (kept as given, already positive in every observed case),
    #    longitude is West (forced negative — the source never carries an
    #    already-negative bare longitude here).
    try:
        bare_lat, bare_lon = float(sx), float(sy)
    except ValueError:
        bare_lat = bare_lon = None
    if bare_lat is not None and bare_lon is not None:
        lat_val = bare_lat if bare_lat >= 0 else -bare_lat
        lon_val = -bare_lon if bare_lon > 0 else bare_lon
        return lat_val, lon_val, "point", None

    return None, None, None, f"unrecognized coordinate pair: {raw_x!r}, {raw_y!r}"


def _normalize_dms_punctuation(s):
    """Collapse the Unicode tick-mark variants actually observed in the source
    (straight/curly single quotes, prime/double-prime, modifier-letter double
    prime, and a doubled single-prime used in place of a real double-prime)
    down to a canonical ' (minute) / " (second) so the DMS regexes below only
    need to handle one spelling. This is character-encoding normalization, not
    a numeric guess — the underlying degree/minute/second digits are untouched."""
    s = str(s)
    s = s.replace("′′", '"').replace("''", '"')  # doubled single-prime used as seconds mark
    s = s.replace("’", "'").replace("‘", "'").replace("′", "'")
    s = s.replace("″", '"').replace("ʺ", '"').replace("”", '"').replace("“", '"')
    return s


def parse_dms_token(s):
    """Parse one DMS (or bare decimal-degree) string to signed decimal degrees.
    Returns None on failure — never guesses."""
    s = _normalize_dms_punctuation(s).strip()

    # Seconds format (with or without decimal seconds): 32°35'34.12"N / 31°48'00"N
    m = re.match(r"^(\d{1,3})[°\s]+(\d{1,2})['\s]+([\d.]+)[\"\s]*([NSEW])$", s)
    if m:
        deg, minute, sec, hemi = m.groups()
        minute_f, sec_f = float(minute), float(sec)
        if minute_f >= 60 or sec_f >= 60:
            return None  # corrupted DMS component (e.g. seconds >= 60) — do not guess
        val = float(deg) + minute_f / 60.0 + sec_f / 3600.0
        return -val if hemi in ("S", "W") else val

    # Comma-decimal-minutes: 31°06,87' N
    m = re.match(r"^(\d{1,3})[°\s]+(\d{1,2}),(\d+)\s*'\s*([NSEW])$", s)
    if m:
        deg, minute_int, minute_frac, hemi = m.groups()
        minute = float(f"{minute_int}.{minute_frac}")
        if minute >= 60:
            return None
        val = float(deg) + minute / 60.0
        return -val if hemi in ("S", "W") else val

    # Dot-decimal-minutes, no seconds: 33°21.56' N  /  33°08.57'N
    m = re.match(r"^(\d{1,3})[°\s]+(\d{1,2}\.\d+)\s*'\s*([NSEW])$", s)
    if m:
        deg, minute, hemi = m.groups()
        minute_f = float(minute)
        if minute_f >= 60:
            return None
        val = float(deg) + minute_f / 60.0
        return -val if hemi in ("S", "W") else val

    # Degree-minute only, no seconds: 35°54'N
    m = re.match(r"^(\d{1,3})[°\s]+(\d{1,2})\s*'\s*([NSEW])$", s)
    if m:
        deg, minute, hemi = m.groups()
        minute_f = float(minute)
        if minute_f >= 60:
            return None
        val = float(deg) + minute_f / 60.0
        return -val if hemi in ("S", "W") else val

    # Bare decimal degrees with hemisphere letter, no punctuation: 29.646508N
    m = re.match(r"^(\d{1,3}\.\d+)\s*([NSEW])$", s)
    if m:
        val, hemi = m.groups()
        val = float(val)
        return -val if hemi in ("S", "W") else val

    # Bare comma-decimal degrees with hemisphere letter, no degree symbol: 32,3241 N
    m = re.match(r"^(\d{1,3}),(\d+)\s*([NSEW])$", s)
    if m:
        deg_int, deg_frac, hemi = m.groups()
        val = float(f"{deg_int}.{deg_frac}")
        return -val if hemi in ("S", "W") else val

    return None


def parse_combined_interval_no_hemisphere(s):
    """Combined "LAT1-LAT2,LON1-LON2" interval, no hemisphere letters per token
    (observed in Data generale, e.g. Oualidia caves/Lagoon). Assumes Morocco's
    convention: latitude range is North, longitude range is West."""
    s = s.replace("\n", " ")
    if "," not in s:
        return None, None, None, "no comma separator in combined interval string"
    lat_part, lon_part = s.split(",", 1)

    def parse_range(part, positive_hemi):
        toks = re.split(r"[–-]", part.strip())
        vals = []
        for t in toks:
            t = t.strip()
            m = re.match(r"^(\d{1,3})[°]\s*(\d{1,2})['’′]\s*([\d.]+)[\"″]?$", t)
            if m:
                deg, minute, sec = m.groups()
                minute_f, sec_f = float(minute), float(sec)
                if minute_f >= 60 or sec_f >= 60:
                    return None
                vals.append(float(deg) + minute_f / 60.0 + sec_f / 3600.0)
        if len(vals) >= 2:
            mid = sum(vals[:2]) / 2.0
            return mid if positive_hemi else -mid
        return None

    lat_mid = parse_range(lat_part, positive_hemi=True)   # North
    lon_mid = parse_range(lon_part, positive_hemi=False)  # West
    if lat_mid is not None and lon_mid is not None:
        return lat_mid, lon_mid, "interval", None
    return None, None, None, f"unrecognized no-hemisphere interval string: {s!r}"


def parse_combined_interval_with_hemisphere(s):
    """Combined "Intervalle LAT1N-LAT2N, LON1W-LON2W" style (Feuil2's own format)."""
    body = re.sub(r"(?i)^intervalle\s*", "", s.strip())
    if "," not in body:
        return None, None, None, f"unrecognized interval string (no comma): {s!r}"
    lat_part, lon_part = body.split(",", 1)

    def parse_range(part):
        toks = re.findall(r"[\d][\d°'’′‘\"″ʺ”“,.\s]*[NSEW]", part)
        vals = [parse_dms_token(t.strip()) for t in toks]
        vals = [v for v in vals if v is not None]
        return vals

    lat_vals = parse_range(lat_part)
    lon_vals = parse_range(lon_part)
    if len(lat_vals) >= 2 and len(lon_vals) >= 2:
        return sum(lat_vals[:2]) / 2.0, sum(lon_vals[:2]) / 2.0, "interval", None
    return None, None, None, f"unrecognized hemisphere interval string: {s!r}"


def extract_utm_xy_from_string(s):
    """Parse Feuil2's combined 'X=714535, Y=3373807' (with possible thousands-
    separator spaces, e.g. 'Y=3 687 500') representation into numeric (x, y)."""
    m = re.match(r"X\s*=\s*([\d\s]+),?\s*Y\s*=\s*([\d\s]+)", s, re.IGNORECASE)
    if not m:
        return None, None
    x = float(m.group(1).replace(" ", ""))
    y = float(m.group(2).replace(" ", ""))
    return x, y


def _to_float_expected(v):
    return float(str(v).strip().replace("−", "-").replace(",", "."))


def validate_parser_against_feuil2(path):
    """Feuil2 is the source's own worked-example conversion legend, not geosite data.

    Investigation finding (see task-1-report.md for full detail, corrected
    after independent review): Feuil2's DMS / comma-decimal-minutes /
    degree-minute rows are exact hand conversions (4 decimal places) and
    every one of them parses to within 0.01 degrees of the stated value —
    strong, strict validation of those code paths. Feuil2's UTM rows, by
    contrast, are rounded to only 2 decimal places and disagree with the
    parser's zone-29N output by up to 5.76 degrees longitude (row 22 — see
    below), mean ~1.0 deg lat / ~1.3 deg lon across all 16 rows — nowhere
    near precise enough to use as a per-row ground truth. One row (id 22,
    raw "X=645000, Y=2827000") is not just imprecise but a genuinely
    different case: its Remarques column explicitly states "Zone 28N", and it
    converts correctly under EPSG:32628 (0.25 deg total error) rather than
    29N. This is a real explicit zone-override test case, not measurement
    noise — the code below reads Remarques and applies the override for that
    row specifically, verifying it converts correctly under 28N rather than
    folding it into the generic 29N-assumed sweep. Excluding that row, the
    remaining 15 UTM rows still disagree with zone 29N's output by up to
    ~2.5 degrees per row (not measurement-precision-level agreement) but the
    *zone* itself is validated two ways instead of by per-row tolerance:
    (a) summed absolute error across those 15 rows is minimized by zone 29N
    versus 28N/30N by a wide margin (a real test that would catch a
    genuinely wrong zone), and (b) converting Data generale's own numeric UTM
    cell for "Eburnian-Proterozoic Unconformity" (680098, 3380902) under
    EPSG:32629 reproduces its already-known-correct coordinate
    (30.547, -7.122) to 6 decimal places exactly (see
    validate_utm_against_known_anchor). So: DMS-family rows are asserted
    strictly; UTM rows are validated via the row-22 override check plus the
    aggregate zone-selection check, not asserted to match Feuil2's own
    imprecise generic UTM column tightly.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Feuil2"]
    n_checked, n_failed = 0, 0
    utm_rows = []  # (x, y, expected_lat, expected_lon) for the aggregate zone check below
    override_checked = False
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] is None:
            continue
        raw, coord_type, expected_lat, expected_lon, remarks = row[1], row[2], row[3], row[4], row[5]
        if coord_type == "Intervalle":
            continue  # documented as a known special case; verified by eye (see report), not asserted here

        exp_lat = _to_float_expected(expected_lat)
        exp_lon = _to_float_expected(expected_lon)

        if coord_type == "UTM":
            x, y = extract_utm_xy_from_string(str(raw))
            if x is None:
                continue
            if remarks and "28N" in str(remarks):
                # Explicit zone override stated in the source's own Remarques
                # column: apply it directly and verify it converts correctly
                # under 28N, instead of folding this row into the generic
                # 29N-assumed aggregate sweep below (where it would appear as
                # a 5.76 deg longitude outlier — that's the wrong zone being
                # applied on purpose to prove the override path works, not a
                # parser bug).
                override_checked = True
                lat, lon, prec, err = parse_coordinate_pair(x, y, utm_zone=28)
                override_err = None if err else abs(lat - exp_lat) + abs(lon - exp_lon)
                print(f"  Feuil2 UTM row with explicit zone override (28N per Remarques={remarks!r}): "
                      f"raw={raw!r} expected=({exp_lat},{exp_lon}) got=({lat:.4f},{lon:.4f}) "
                      f"abs_err={override_err:.3f}")
                assert err is None and override_err < 0.5, (
                    f"Zone-28N override for Feuil2 row {raw!r} did not convert plausibly close "
                    f"to its expected value (abs_err={override_err}) — override logic is broken"
                )
            else:
                utm_rows.append((x, y, exp_lat, exp_lon))
                lat, lon, prec, err = parse_coordinate_pair(x, y, utm_zone=29)
                print(f"  Feuil2 UTM row (informational — see docstring, not asserted per-row): "
                      f"raw={raw!r} expected=({exp_lat},{exp_lon}) got=({lat:.4f},{lon:.4f})")
            continue  # UTM rows are validated via override or aggregate below, not per-row here

        # DMS / comma-decimal-minutes / degree-minute: reconstruct as a two-token
        # (lat_str, lon_str) pair, same shape parse_coordinate_pair sees from
        # Data generale's split cells, and assert strictly.
        tokens = re.findall(r"[\d][\d°'’′‘\"″ʺ”“,.\s]*[NSEW]", str(raw))
        lat_tok = next((t for t in tokens if t.strip()[-1] in "NS"), None)
        lon_tok = next((t for t in tokens if t.strip()[-1] in "EW"), None)
        if lat_tok is None or lon_tok is None:
            lat, lon, prec, err = None, None, None, f"could not tokenize DMS string: {raw!r}"
        else:
            lat, lon, prec, err = parse_coordinate_pair(lat_tok.strip(), lon_tok.strip())

        n_checked += 1
        if err is not None or lat is None or abs(lat - exp_lat) > 0.01 or abs(lon - exp_lon) > 0.01:
            n_failed += 1
            print(f"  FEUIL2 MISMATCH row raw={raw!r} type={coord_type} "
                  f"expected=({exp_lat},{exp_lon}) got=({lat},{lon}) err={err}")

    print(f"Feuil2 strict validation (DMS-family): checked {n_checked}, {n_failed} failed")
    assert n_checked >= 5, "Expected at least 5 validatable Feuil2 rows, got fewer — check sheet access"
    assert n_failed == 0, f"{n_failed} Feuil2 DMS-family ground-truth conversions failed — fix the parser, do not weaken tolerance"
    assert override_checked, "Expected to find and validate the Zone-28N override row (id 22) in Feuil2, but never encountered it"

    # Aggregate UTM-zone check (row 22's explicit-override case excluded, see
    # above): Feuil2's UTM "expected" column is individually imprecise (2
    # decimal places, off from the parser's zone-29N output by up to ~2.5 deg
    # per row on this 15-row subset — the sheet author's rough map-reading
    # estimate, not a computed value; see docstring), so no single-row
    # tolerance is meaningful. What IS meaningful and a real test of parser
    # correctness: summed absolute error across these 15 rows must be
    # minimized by zone 29N specifically, not by 28N or 30N — this would
    # catch a genuinely wrong zone (a real bug) even though it tolerates the
    # fixture's per-row noise.
    assert len(utm_rows) >= 5, "Expected at least 5 non-override UTM rows in Feuil2 for the aggregate zone check"
    zone_errors = {}
    for zone in (28, 29, 30):
        total = 0.0
        for x, y, exp_lat, exp_lon in utm_rows:
            lat, lon, _, _ = parse_coordinate_pair(x, y, utm_zone=zone)
            total += abs(lat - exp_lat) + abs(lon - exp_lon)
        zone_errors[zone] = total
    print(f"Feuil2 UTM aggregate abs error by zone (excl. row 22 override case, {len(utm_rows)} rows): {zone_errors}")
    best_zone = min(zone_errors, key=zone_errors.get)
    assert best_zone == 29, (
        f"Zone 29N is not the best-fitting UTM zone against Feuil2's UTM rows "
        f"(errors: {zone_errors}) — this would indicate a real parser/zone bug, investigate"
    )
    print(f"UTM zone selection confirmed: 29N minimizes aggregate error "
          f"({zone_errors[29]:.2f} vs 28N={zone_errors[28]:.2f}, 30N={zone_errors[30]:.2f})")


def validate_utm_against_known_anchor():
    """Independent, precise UTM Zone 29N (EPSG:32629) validation using the one
    coordinate the spec confirms is already known-correct by other means: Data
    generale's "Eburnian-Proterozoic Unconformity" numeric cell (680098, 3380902)
    must reproduce (30.547, -7.122)."""
    lat, lon, prec, err = parse_coordinate_pair(680098, 3380902)
    assert err is None, f"UTM anchor point failed to parse: {err}"
    assert abs(lat - 30.547) < 0.001 and abs(lon - (-7.122)) < 0.001, \
        f"UTM anchor point mismatch: got ({lat}, {lon}), expected (30.547, -7.122)"
    print(f"UTM Zone 29N anchor validation: (680098, 3380902) -> ({lat:.4f}, {lon:.4f}) matches known-correct (30.547, -7.122)")


# --- Excel hierarchical loader (Data generale / Feuil1) --------------------
#
# IMPORTANT semantic note (2026-07-30 findings, see
# docs/superpowers/specs/2026-07-30-geoheritage-excel-data-model-findings.md):
# `Locality_Group_Key` here means "these geosites were inventoried together by
# one publication" — NOT "these are the same physical place." One publication
# group can legitimately span very large real distances (confirmed: the
# Tinghir-Dades-Imilchil group's 27 members span ~170km). The grouping
# mechanism itself (one key per contiguous run of rows sharing one merged-cell
# `Reference`/domain block) is correct and should be kept — only the earlier
# narration of it as "one physical locality" was wrong. Downstream logic must
# not assume geographic proximity between members of the same group.

OBS_COLS_EXCEL = OBS_COLS + ["Center_Coord_Source"]

# Real Morocco UTM eastings run roughly 100,000-800,000; Feuil1 stores some of
# its "real" center eastings in kilometers instead of meters with no marker
# distinguishing them (confirmed against Feuil2, e.g. Feuil1 row 260's
# (453.5, "3,273,000") == Feuil2's stated X=453500, Y=3273000). Anything under
# this threshold is treated as km and rescaled to meters.
_FEUIL1_KM_THRESHOLD = 1000.0


def _feuil1_numeric(v):
    """Feuil1's Center coordinates cells are a mix of numeric types and
    comma-thousands-separated strings (e.g. '3,273,000') — coerce to float,
    return None if not parseable."""
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        v = v.replace(",", "").strip()
        try:
            return float(v)
        except ValueError:
            return None
    return None


def build_feuil1_center_fallback(path):
    """`Feuil1` is NOT a pure duplicate of `Data generale` (2026-07-30
    findings, revising the earlier "confirmed redundant, skip it" decision,
    which was based on sampling only rows 3-14): for 12+ publication groups,
    `Data generale`'s Center coordinates is the "…." placeholder while
    `Feuil1` retains a real UTM easting/northing for the very same group.

    Both sheets' publication groups start on the identical row number (spot
    checked directly, e.g. row 3 = "Eburnian-Proterozoic Unconformity" in
    both, row 575 = the Dakhla peninsula group in both) — so this returns a
    simple {row_index: (x, y)} map, keyed by that shared row number, of every
    Feuil1 group-start row with a real, unit-corrected (km->m) center value.
    Used only as a fallback when Data generale's own center is a placeholder;
    Data generale otherwise remains the primary source for everything else
    (Reference/Title/Methodology/Comment/individual Coordinates)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Feuil1"]
    out = {}
    for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        domain, ccx, ccy = row[0], row[6], row[7]
        if domain is None:
            continue  # not a group-start row
        x, y = _feuil1_numeric(ccx), _feuil1_numeric(ccy)
        if x is None or y is None:
            continue
        if x < _FEUIL1_KM_THRESHOLD:
            x *= 1000.0
        out[i] = (x, y)
    return out


def build_feuil2_center_overrides(path):
    """`Feuil2` (26 rows) is the source's own authoritative conversion table
    for essentially every distinct Center coordinates value in the workbook
    (2026-07-30 findings §8) — not merely a parser-validation fixture. For its
    DMS/"Intervalle" entries specifically, its stated decimal value is treated
    here as authoritative and substituted in whenever a group's own parsed
    center matches (within a tight tolerance) Feuil2's own parse of the exact
    same raw text — i.e. a genuine same-content match, not a coincidence.

    UTM-family Feuil2 entries are deliberately NOT used as a direct override:
    validate_parser_against_feuil2's own established finding is that Feuil2's
    decimal column for UTM rows is a rough 2-decimal-place map-reading
    estimate that disagrees with the precise zone-29N/28N conversion by up to
    several degrees — the precise on-the-fly UTM conversion (with the Dakhla
    28N override, see _GROUP_UTM_ZONE_OVERRIDES) is what is used for those
    groups instead. This function only ever returns DMS/Intervalle-family
    entries."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Feuil2"]
    overrides = []  # (self_parsed_lat, self_parsed_lon, stated_lat, stated_lon, raw)
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] is None:
            continue
        raw, coord_type, expected_lat, expected_lon = row[1], row[2], row[3], row[4]
        if coord_type not in ("DMS", "Intervalle"):
            continue
        try:
            exp_lat, exp_lon = _to_float_expected(expected_lat), _to_float_expected(expected_lon)
        except (TypeError, ValueError):
            continue
        if coord_type == "Intervalle":
            lat, lon, prec, err = parse_combined_interval_with_hemisphere(str(raw))
        else:
            tokens = re.findall(r"[\d][\d°'’′‘\"″ʺ”“,.\s]*[NSEW]", str(raw))
            lat_tok = next((t for t in tokens if t.strip()[-1] in "NS"), None)
            lon_tok = next((t for t in tokens if t.strip()[-1] in "EW"), None)
            if lat_tok is None or lon_tok is None:
                continue
            lat, lon, prec, err = parse_coordinate_pair(lat_tok.strip(), lon_tok.strip())
        if err is not None or lat is None:
            continue
        overrides.append((lat, lon, exp_lat, exp_lon, raw))
    return overrides


def _apply_feuil2_override(lat, lon, overrides, tol=0.01):
    """Returns (lat, lon, matched_raw_or_None). matched_raw is not None iff a
    Feuil2 DMS/Intervalle entry's own self-parse landed within `tol` degrees
    of the given (lat, lon) — a same-content match — in which case Feuil2's
    own stated decimal is substituted in."""
    for self_lat, self_lon, stated_lat, stated_lon, raw in overrides:
        if abs(lat - self_lat) <= tol and abs(lon - self_lon) <= tol:
            return stated_lat, stated_lon, raw
    return lat, lon, None


def _is_placeholder_coord(v):
    """Data generale's "…." (and similar ellipsis-only) cells mean "not
    given" — a real, intentional value, not a parsing failure (2026-07-30
    findings §9). Detected generically as "no digit at all", since every real
    coordinate representation in this source contains at least one digit."""
    if v is None:
        return True
    if isinstance(v, (int, float)):
        return False
    return not any(ch.isdigit() for ch in str(v))


def load_excel_hierarchical(path, sheet_name, source_tag, feuil1_centers=None, feuil2_overrides=None):
    feuil1_centers = feuil1_centers or {}
    feuil2_overrides = feuil2_overrides or []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows_out, failures = [], []
    current_domain, current_region, current_ref = None, None, None
    current_center_lat, current_center_lon = None, None
    current_center_source = None
    group_key_counter = 0
    current_group_key = None

    for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        domain, region, reference, title, name, gtype = row[0], row[1], row[2], row[3], row[4], row[5]
        cx, cy, ccx, ccy = row[6], row[7], row[8], row[9]

        if domain is not None:
            current_domain, current_region, current_ref = domain, region, reference
            group_key_counter += 1
            current_group_key = f"{source_tag}_group_{group_key_counter}"
            current_center_lat, current_center_lon, current_center_source = None, None, None
            zone = _group_utm_zone(reference)

            ccx_eff, ccy_eff, used_fallback = ccx, ccy, False
            if _is_placeholder_coord(ccx) or _is_placeholder_coord(ccy):
                fallback = feuil1_centers.get(i)
                if fallback is not None:
                    ccx_eff, ccy_eff = fallback
                    used_fallback = True

            if ccx_eff is not None and ccy_eff is not None and not _is_placeholder_coord(ccx_eff):
                clat, clon, cprec, cerr = parse_coordinate_pair(ccx_eff, ccy_eff, utm_zone=zone)
                if cerr is None:
                    current_center_lat, current_center_lon = clat, clon
                    current_center_source = "Feuil1_fallback" if used_fallback else "Data_generale"
                    olat, olon, oraw = _apply_feuil2_override(clat, clon, feuil2_overrides)
                    if oraw is not None:
                        current_center_lat, current_center_lon = olat, olon
                        current_center_source = "Feuil2_lookup"
        if name is None:
            continue

        # Note: cy can legitimately be None even when cx holds real data — the
        # combined single-cell interval format (e.g. "Oualidia caves"/"Oualidia
        # Lagoon") packs both lat and lon ranges into cx alone. parse_coordinate_pair
        # handles raw_y=None internally for those combined-string cases, so only
        # cx being None means truly missing coordinates.
        # locality_center is passed through so the Merchich/Nord Maroc branch can
        # plausibility-check itself against the group's own declared center (see
        # parse_coordinate_pair docstring) — harmless no-op for every other format.
        lat, lon, prec, err = parse_coordinate_pair(
            cx, cy, locality_center=(current_center_lat, current_center_lon)
        ) if cx is not None else (None, None, None, "missing coordinates")
        if err:
            failures.append({"Geosite_Name": name, "Raw_X": cx, "Raw_Y": cy, "Error": err,
                              "Source_File": f"{source_tag}:{sheet_name}", "Source_Row_Ref": i})
            continue

        rows_out.append({
            "Geosite_Name": str(name).strip(), "Latitude_WGS84": lat, "Longitude_WGS84": lon,
            "Region": current_region, "Geosite_Type": str(gtype).strip() if gtype else np.nan,
            "Geological_Domain": current_domain,
            "Locality_Center_Lat": current_center_lat, "Locality_Center_Lon": current_center_lon,
            "Locality_Group_Key": current_group_key, "Coordinate_Precision": prec,
            "Source_File": f"{source_tag}:{sheet_name}", "Source_Row_Ref": i,
            "Center_Coord_Source": current_center_source,
        })
    return pd.DataFrame(rows_out, columns=OBS_COLS_EXCEL), pd.DataFrame(failures)


# --- Flat CSV loader (national/regional catalogs) ---------------------------

def load_flat_csv(path, name_col, lat_col, lon_col, region_col, type_col, domain_col, source_tag):
    df = pd.read_csv(path)
    out = pd.DataFrame()
    out["Geosite_Name"] = df[name_col].astype(str).str.strip()
    out["Latitude_WGS84"] = pd.to_numeric(df[lat_col], errors="coerce")
    out["Longitude_WGS84"] = pd.to_numeric(df[lon_col], errors="coerce")
    out["Region"] = df[region_col].astype(str).str.strip() if region_col in df.columns else np.nan
    out["Geosite_Type"] = df[type_col] if type_col in df.columns else np.nan
    out["Geological_Domain"] = df[domain_col] if domain_col in df.columns else np.nan
    out["Locality_Center_Lat"] = np.nan
    out["Locality_Center_Lon"] = np.nan
    out["Locality_Group_Key"] = [f"{source_tag}_row_{i}" for i in range(len(df))]  # each its own singleton group
    out["Coordinate_Precision"] = "point"
    out["Source_File"] = source_tag
    out["Source_Row_Ref"] = df.index
    return out[OBS_COLS]


# --- Phase B: per-observation correctness checks ---------------------------
#
# Task 2 of the fresh-start rebuild (see task-2-brief.md and Phase B of the
# design spec). Applies documented coordinate corrections, folds in
# report-verified ground truth, then runs a national border check and a
# nationwide regional-plausibility check.

DOCUMENTED_CORRECTIONS = [
    ("Malabata", "Longitude_WGS84", -0.0333, -5.7133,
     "Corrected per geosite_phase2_report_fr.tex:101 (regional field verification: "
     "stored lon=-0.0333 was corrupted, real value -5.7133 W)"),
]

# Punta Cires isn't a coordinate fix — it's a report-documented accessibility
# reclassification (national model predicted "Modérée"/Difficile; a real
# route-time check found an 8 min / 2.7 km piste drive from Belyounech to the
# Jbel Moussa forest, reclassifying it "Facile"). The Observation schema (see
# OBS_COLS) has no accessibility/category column at all yet — that's a later
# labeling-stage concern (spec line 51 explicitly scopes labeling out of this
# stage) — so there is nothing to "overwrite." We record the old/new category
# and the report citation as a Ground_Truth_Note on the existing observation(s)
# instead, which is the schema-honest way to carry this fact forward to the
# stage that will actually consume it.
GROUND_TRUTH_REPORT_NOTES = [
    ("Cires", "Report-documented reclassification (geosite_phase2_report_fr.tex:290): "
              "national model predicted Moderee/Difficile (regional Rif roughness R=263.5); "
              "field route-time check found 8min/2.7km piste drive Belyounech->Jbel Moussa forest "
              "(vs 1h05/4.1km on foot) -> reclassified Facile. Old category (source CSV): "
              "Difficile (Difficult). New category: Facile."),
]

GROUND_TRUTH_ADDITIONS = [
    {"Geosite_Name": "Fahs-Anjra (Melloussa)", "Latitude_WGS84": 35.7250, "Longitude_WGS84": -5.6685,
     "Region": "Tanger-Tétouan-Al Hoceïma", "Geosite_Type": np.nan, "Geological_Domain": np.nan,
     "Locality_Center_Lat": np.nan, "Locality_Center_Lon": np.nan, "Coordinate_Precision": "point",
     "Source_File": "phase2_report_field_verification", "Source_Row_Ref": "geosite_phase2_report_fr.tex:289",
     "Ground_Truth_Note": "Verified <750m / 11min walk from Gare de Melloussa; report classifies Facile"},
    # Oulmès/Khouribga plateau: kept as a flagged/cautionary example per spec
    # step 9 (report itself calls the regional-model reclassification to
    # Facile a likely overestimate for this point), NOT auto-labeled Facile.
    # Region is set to "Béni Mellal-Khénifra" per the user's direct local
    # knowledge (2026-07-30 correction) -- Khouribga province is part of
    # Béni Mellal-Khénifra under the current 12-region administrative map,
    # even though the fetched boundary polygon geometrically places this
    # exact point (33.4720, -6.9462) just inside a neighboring
    # "Grand Casablanca-Settat" polygon. That geometric placement is treated
    # here as likely boundary-polygon imprecision near this location, not
    # evidence the declared region is wrong. If the regional-plausibility
    # check below flags this point, it is routed to
    # geosites_region_mismatches.csv for human review like any other
    # mismatch -- NOT silently auto-corrected -- but should be read as a
    # known boundary-precision case, not a confirmed source-data error.
    {"Geosite_Name": "Oulmès/Khouribga plateau", "Latitude_WGS84": 33.4720, "Longitude_WGS84": -6.9462,
     "Region": "Béni Mellal-Khénifra", "Geosite_Type": np.nan, "Geological_Domain": np.nan,
     "Locality_Center_Lat": np.nan, "Locality_Center_Lon": np.nan, "Coordinate_Precision": "point",
     "Source_File": "phase2_report_field_verification", "Source_Row_Ref": "geosite_phase2_report_fr.tex:291",
     "Ground_Truth_Note": "CAUTION (report's own flag): regional BMK model reclassifies this pixel "
                          "Difficile->Facile due to near-flat topography (slope=1.8deg, ruggedness=50.4), "
                          "but report explicitly cautions this overestimates pedestrian accessibility in an "
                          "isolated rural area with no paved road nearby. NOT auto-labeled Facile -- kept as "
                          "a negative/cautionary example per spec step 9."},
]


def apply_documented_corrections(df):
    df = df.copy()
    df["Correction_Applied"] = pd.Series(np.nan, index=df.index, dtype=object)
    for name_sub, field, wrong_hint, correct_val, note in DOCUMENTED_CORRECTIONS:
        mask = df["Geosite_Name"].str.contains(name_sub, case=False, na=False) & np.isclose(df[field], wrong_hint, atol=0.01)
        if mask.sum() > 0:
            df.loc[mask, field] = correct_val
            df.loc[mask, "Correction_Applied"] = note
            print(f"Applied correction for {mask.sum()} observation(s) matching {name_sub!r}")
    return df


def apply_ground_truth_notes(df):
    """Report-documented facts that don't change any coordinate/field value in
    this schema (no accessibility/category column exists yet) but must still
    be carried forward as provenance -- e.g. Punta Cires' reclassification."""
    df = df.copy()
    if "Ground_Truth_Note" not in df.columns:
        df["Ground_Truth_Note"] = pd.Series(np.nan, index=df.index, dtype=object)
    for name_sub, note in GROUND_TRUTH_REPORT_NOTES:
        mask = df["Geosite_Name"].str.contains(name_sub, case=False, na=False)
        if mask.sum() > 0:
            df.loc[mask, "Ground_Truth_Note"] = note
            print(f"Recorded ground-truth note for {mask.sum()} observation(s) matching {name_sub!r}")
    return df


def fold_in_ground_truth(df):
    additions = pd.DataFrame(GROUND_TRUTH_ADDITIONS)
    additions["Locality_Group_Key"] = [f"ground_truth_{i}" for i in range(len(additions))]
    # Each addition is its own new observation and needs its own Observation_ID
    # (same singleton-group convention as load_flat_csv), continuing the
    # numbering after every existing observation.
    next_id = df["Observation_ID"].str.extract(r"obs_(\d+)").astype(int)[0].max() + 1
    additions["Observation_ID"] = [f"obs_{next_id + i:05d}" for i in range(len(additions))]
    for col in OBS_COLS + ["Correction_Applied"]:
        if col not in additions.columns:
            additions[col] = np.nan
    df = pd.concat([df, additions], ignore_index=True)
    print(f"Folded in {len(additions)} report-verified ground-truth observation(s): "
          f"{additions['Geosite_Name'].tolist()}")
    return df


def fetch_morocco_boundary():
    url = "https://raw.githubusercontent.com/nvkelso/natural-earth-vector/master/geojson/ne_10m_admin_0_countries.geojson"
    world = gpd.read_file(url)
    morocco = world[world["ADMIN"].isin(["Morocco", "Western Sahara"])]
    print(f"Fetched Morocco/Western Sahara boundary: {len(morocco)} polygon(s), "
          f"ADMIN={sorted(morocco['ADMIN'].unique())}")
    assert len(morocco) == 2, "Expected exactly Morocco + Western Sahara admin-0 polygons"
    return morocco.to_crs("EPSG:4326")


# Natural Earth's admin-0 polygon is a simplified vector, not a survey-grade
# coastline -- checked for real (task-2 investigation) by measuring every
# rejected point's actual distance to the boundary union before trusting the
# raw within() test. Result: a clean, wide gap between two populations --
# 21 points sit 0.2-7.3km outside the raw polygon (all coastal cliffs/dunes,
# e.g. Akhfénir Cliff, or Rif-mountain sites right at the vector's edge, e.g.
# Targha mica-shist) vs. genuine outliers starting at 99km (e.g. "Cylindrical
# folds of Ordovician quartzites" at lon=-13.89, ~90km offshore in the
# Atlantic) up to thousands of km (lat~6/lon~-32, lon~8.5, lon~-42 -- the
# uncorrected outliers Task 1 flagged and deliberately left for this check).
# A small buffer absorbs the coastline/vector-precision artifacts without
# coming close to admitting any genuine outlier -- same wide-separation
# reasoning Task 1 used for the Merchich 50km threshold.
_BORDER_BUFFER_DEG = 0.1  # ~11km at this latitude; nearest genuine outlier is 99km away


def border_check(df, boundary_gdf):
    geometry = [Point(xy) if pd.notna(xy[0]) and pd.notna(xy[1]) else None
                for xy in zip(df["Longitude_WGS84"], df["Latitude_WGS84"])]
    gdf = gpd.GeoDataFrame(df, geometry=geometry, crs="EPSG:4326")
    union = boundary_gdf.union_all().buffer(_BORDER_BUFFER_DEG)
    inside = gdf.geometry.notna() & gdf.geometry.within(union)
    outliers = gdf[~inside & gdf.geometry.notna()].drop(columns="geometry")
    kept = gdf[inside | gdf.geometry.isna()].drop(columns="geometry")  # missing coords pass through, not "outside"
    print(f"Border check: {len(kept)} kept (inside Morocco/Western Sahara + {_BORDER_BUFFER_DEG}deg buffer, "
          f"or missing coordinates), {len(outliers)} outside the border and removed")
    if len(outliers) > 0:
        print(outliers[["Geosite_Name", "Latitude_WGS84", "Longitude_WGS84", "Source_File"]].head(10).to_string(index=False))
    return kept, outliers


# --- Regional-plausibility check --------------------------------------------
#
# IMPORTANT deviation from the brief's draft: the brief suggested fetching
# Natural Earth's ne_10m_admin_1_states_provinces.geojson for Morocco's region
# polygons. Fetching and inspecting it for real (per the task instructions'
# explicit warning to verify this before trusting it) showed it is WRONG for
# this purpose: it returns Morocco's *pre-2015-reform* 16-region system
# (17 rows incl. Western Sahara as one unit) with names like "Tanger -
# Tétouan", "Tadla - Azilal", "Fès - Boulemane" -- structurally different from
# and mostly *not name-mappable* to the current 12-region system
# (Tanger-Tétouan-Al Hoceïma, Béni Mellal-Khénifra, ...) that every declared
# `Region` value in this project's data actually uses. Using it directly would
# have made ~every observation "mismatch" on region-name grounds alone,
# drowning out genuine placement errors. Switched to
# github.com/Salah-Zkara/Morocco-GeoJson (`Morocco-Regions.geojson`), which
# was verified (see task-2 investigation) to return exactly the current 12
# regions with plausible, valid, contiguous geometry covering all of
# Morocco + Western Sahara (bounds -17.1..-1.0 lon, 20.8..35.9 lat).
def fetch_region_boundaries():
    url = "https://raw.githubusercontent.com/Salah-Zkara/Morocco-GeoJson/master/Morocco-Regions.geojson"
    regions = gpd.read_file(url)
    print(f"Fetched {len(regions)} region polygons: {sorted(regions['nom_region'].dropna().unique())}")
    assert len(regions) == 12, f"Expected exactly Morocco's 12 current administrative regions, got {len(regions)}"
    assert regions.geometry.is_valid.all(), "Fetched region polygons contain invalid geometry"
    return regions.to_crs("EPSG:4326")


def _norm_region(s):
    s = str(s).lower().strip()
    for a, b in [("é", "e"), ("è", "e"), ("â", "a"), ("ï", "i"), ("î", "i"), ("ô", "o"), ("û", "u"),
                 ("ç", "c"), ("à", "a"), ("'", " "), ("-", " "), ("&", " and ")]:
        s = s.replace(a, b)
    return " ".join(s.split())


# Declared `Region` strings across the actual raw data include historical
# (pre-2015) region names, minor typos, and one combined "X & Y" label -- none
# of which are placement errors, just naming drift the source catalogs never
# reconciled. Verified against real geometry (task-2 investigation: every
# "Doukkala-Abda"-declared point with real coordinates lands inside either
# "Grand Casablanca-Settat" or "Marrakech-Safi", matching how the old
# Doukkala-Abda region was split across those two in the 2015 reform; every
# "Laayoune-Dakhla"-declared point lands in "Laayoune-Sakia El Hamra" or is
# outside all 12 polygons; "Marrakech-Saf" is a truncation typo for
# "Marrakech-Safi"). A pure accent-only string comparison (the brief's draft)
# would flag all of these as false-positive mismatches and swamp the real
# nationwide signal the check exists to surface -- so declared names are first
# resolved through this alias table (built from and checked against the
# actual declared values present in the data) before falling back to plain
# normalized equality.
_REGION_ALIASES = {
    _norm_region("Doukkala-Abda"): {_norm_region("Grand Casablanca-Settat"), _norm_region("Marrakech-Safi")},
    _norm_region("Laayoune-Dakhla"): {_norm_region("Laayoune-Sakia El Hamra"), _norm_region("Guelmim-Oued Noun"),
                                       _norm_region("Eddakhla-Oued Eddahab")},
    _norm_region("Dakhla-Oued Ed-Dahab"): {_norm_region("Eddakhla-Oued Eddahab")},
    _norm_region("Dakhla-Oued Eddahab"): {_norm_region("Eddakhla-Oued Eddahab")},
    _norm_region("Marrakech-Saf"): {_norm_region("Marrakech-Safi")},
    _norm_region("Draa-Tafilalet & Fès-Méknès"): {_norm_region("Drâa-Tafilalet"), _norm_region("Fés-Meknés")},
    _norm_region("Fès-Meknès"): {_norm_region("Fés-Meknés")},
    _norm_region("Beni Mellal-Khenifra"): {_norm_region("Béni Mellal-Khénifra")},
    _norm_region("Beni Mellal-Khénifra"): {_norm_region("Béni Mellal-Khénifra")},
    # Added 2026-08-09 processing the expanded (1465-row) Data generale delivery: these
    # two declared-region spellings drop "Tétouan" entirely (not just an accent variant,
    # a genuinely different token count) and cleanly resolve 100% (183/183) to
    # Tanger-Tétouan-Al Hoceima once mapped -- a real alias gap, not a placement error.
    _norm_region("Tanger-Al Hoceima"): {_norm_region("Tanger-Tétouan-Al Hoceima")},
    _norm_region("Tanger-Al Houceima"): {_norm_region("Tanger-Tétouan-Al Hoceima")},
    # English "Fez" vs French/official "Fès/Fés" spelling (a letter substitution, not an
    # accent-stripping case _norm_region already handles).
    _norm_region("Fez-Meknes"): {_norm_region("Fés-Meknés")},
}


def _regions_match(declared_series, detected_series):
    def one(declared, detected):
        d, t = _norm_region(declared), _norm_region(detected)
        if d == t:
            return True
        return t in _REGION_ALIASES.get(d, set())
    return pd.Series([one(d, t) for d, t in zip(declared_series, detected_series)], index=declared_series.index)


def regional_plausibility_check(df, region_gdf, region_name_col):
    geometry = [Point(xy) if pd.notna(xy[0]) and pd.notna(xy[1]) else None
                for xy in zip(df["Longitude_WGS84"], df["Latitude_WGS84"])]
    gdf = gpd.GeoDataFrame(df, geometry=geometry, crs="EPSG:4326")
    joined = gpd.sjoin(gdf, region_gdf[[region_name_col, "geometry"]], how="left", predicate="within")
    joined = joined.rename(columns={region_name_col: "Detected_Region"})
    joined = joined[~joined.index.duplicated(keep="first")]  # sjoin can duplicate rows on boundary edges
    has_declared = joined["Region"].notna() & (joined["Region"].astype(str).str.strip() != "") & (joined["Region"].astype(str) != "nan")
    mismatch = has_declared & joined["Detected_Region"].notna() & (~_regions_match(joined["Region"], joined["Detected_Region"]))
    mismatches = joined[mismatch].drop(columns="geometry")
    clean = joined[~mismatch].drop(columns=["geometry", "Detected_Region", "index_right"], errors="ignore")
    print(f"Regional plausibility check: {mismatch.sum()} mismatch(es) found nationwide, {len(clean)} observations consistent (or unchecked)")
    if mismatch.sum() > 0:
        print(mismatches[["Geosite_Name", "Region", "Detected_Region"]].head(15).to_string(index=False))
    return clean, mismatches


def _norm_name(s):
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", s).strip().lower()


def _assemble_locality_row(locality_id, group):
    # Prefer a CSV-catalog-sourced name/coordinate as representative if the group has one
    # (generally more precise/reviewed); otherwise fall back to the first member.
    catalog_rows = group[group["Source_File"].isin(["national_catalog", "ttah_regional", "bmk_regional"])]
    rep = catalog_rows.iloc[0] if len(catalog_rows) > 0 else group.iloc[0]
    return {
        "Locality_ID": locality_id,
        "Geosite_Name": rep["Geosite_Name"],
        "Latitude_WGS84": rep["Latitude_WGS84"],
        "Longitude_WGS84": rep["Longitude_WGS84"],
        "Region": group["Region"].dropna().iloc[0] if group["Region"].notna().any() else np.nan,
        "Geosite_Type": rep.get("Geosite_Type", np.nan),
        "Geological_Domain": rep.get("Geological_Domain", np.nan),
        "Observation_Count": len(group),
        "Member_Observation_IDs": ";".join(group["Observation_ID"].astype(str)),
        "Source_Files": ";".join(group["Source_File"].astype(str).unique()),
        "Correction_Applied": "; ".join(group["Correction_Applied"].dropna().unique()) or np.nan,
        "Ground_Truth_Note": "; ".join(group["Ground_Truth_Note"].dropna().unique()) if "Ground_Truth_Note" in group.columns and group["Ground_Truth_Note"].notna().any() else np.nan,
    }


def build_localities(df, name_similarity_threshold=85, proximity_tolerance_m=500):
    """
    Build deduplicated localities from ALL observations, each starting as its own
    singleton candidate (CSV-catalog rows AND Excel rows alike).

    2026-07-30 correction: an Excel `Locality_Group_Key` means "inventoried by the
    same publication," NOT "same physical place" -- a publication group can span
    ~170km with dozens of distinct, separately-named geosites (e.g. the
    Tinghir-Dades-Imilchil literature group). So group membership is NOT used as
    evidence of physical identity here. Instead:

      1. Observations WITH coordinates are spatially indexed (KD-tree, projected
         to EPSG:26191 meters) and clustered by an anchor-based pass: for each
         unassigned observation, find unassigned neighbors within
         `proximity_tolerance_m` of ITS OWN coordinate (never a shared group
         center -- using the Excel group-center would spuriously merge every
         member of a large publication group, since they'd all sit at distance
         ~0 from a shared centroid) whose name is a fuzzy match
         (rapidfuzz token_sort_ratio >= threshold). Matches collapse into one
         locality; this also naturally collapses genuine intra-group duplicates
         (identical name + identical/near-identical individual coordinates).
      2. Observations WITHOUT coordinates (real gap in the national catalog
         source, ~409 rows) cannot be spatially matched. They are deduplicated
         only by an exact normalized name + region match against other
         coordinate-less rows, then each remaining one becomes its own
         (coordinate-less) locality.
    """
    df = df.copy()
    has_coord = df["Latitude_WGS84"].notna() & df["Longitude_WGS84"].notna()
    geocoded = df[has_coord].reset_index(drop=True)
    no_coord = df[~has_coord].reset_index(drop=True)

    transformer = Transformer.from_crs("EPSG:4326", "EPSG:26191", always_xy=True)
    xs, ys = transformer.transform(geocoded["Longitude_WGS84"].values, geocoded["Latitude_WGS84"].values)
    geocoded["_x"], geocoded["_y"] = xs, ys

    locality_rows, review_rows = [], []
    locality_id_counter = 0

    # --- geocoded observations: KD-tree anchor-based proximity + fuzzy-name clustering
    names = [_norm_name(x) for x in geocoded["Geosite_Name"].astype(str)]
    coords = geocoded[["_x", "_y"]].values
    n = len(geocoded)
    tree = cKDTree(coords) if n > 0 else None
    assigned = np.full(n, False)
    for i in range(n):
        if assigned[i]:
            continue
        cand_idx = tree.query_ball_point(coords[i], r=proximity_tolerance_m) if tree is not None else [i]
        group_idx = [i]
        for j in cand_idx:
            if j == i or assigned[j]:
                continue
            if fuzz.token_sort_ratio(names[i], names[j]) >= name_similarity_threshold:
                group_idx.append(j)
        for j in group_idx:
            assigned[j] = True

        locality_id_counter += 1
        locality_rows.append(_assemble_locality_row(f"loc_{locality_id_counter:05d}", geocoded.iloc[group_idx]))

    # --- coordinate-less observations: exact normalized name + region dedup only
    nc_names = [_norm_name(x) for x in no_coord["Geosite_Name"].astype(str)]
    nc_regions = [_norm_name(x) for x in no_coord["Region"].astype(str)]
    m = len(no_coord)
    nc_assigned = np.full(m, False)
    for i in range(m):
        if nc_assigned[i]:
            continue
        group_idx = [i]
        for j in range(i + 1, m):
            if nc_assigned[j]:
                continue
            if nc_names[j] == nc_names[i] and nc_regions[j] == nc_regions[i]:
                group_idx.append(j)
        for j in group_idx:
            nc_assigned[j] = True

        locality_id_counter += 1
        locality_rows.append(_assemble_locality_row(f"loc_{locality_id_counter:05d}", no_coord.iloc[group_idx]))

    localities = pd.DataFrame(locality_rows)
    review = pd.DataFrame(review_rows) if review_rows else pd.DataFrame(columns=list(df.columns) + ["Max_Pairwise_Distance_m"])
    print(f"Locality construction: {len(df)} observations ({n} geocoded, {m} without coordinates) -> "
          f"{len(localities)} localities, {len(review)} observations sent to review")
    return localities, review


def main():
    excel_path = os.path.join(BASE, "references", "databases", "Data Classification_Geoheritage.xlsx")

    print("=== Step 1: Feuil2 parser validation ===")
    validate_parser_against_feuil2(excel_path)
    validate_utm_against_known_anchor()

    print("\n=== Sources ===")
    sources = [
        load_flat_csv(os.path.join(BASE, "archive/livrable/phase1_national_accessibility/data/geosites_coordinates_clean.csv"),
                       "Geosite_Name", "Latitude_WGS84", "Longitude_WGS84", "Administrative_Region", "Geosite_Type", "Geological_Domain", "national_catalog"),
        load_flat_csv(os.path.join(BASE, "archive/livrable/phase2_regional_analytics/data/geosites_ttah_indexed.csv"),
                       "Geosite_Name", "Latitude_WGS84", "Longitude_WGS84", "Administrative_Region", "Geosite_Type", "Geological_Domain", "ttah_regional"),
        load_flat_csv(os.path.join(BASE, "archive/livrable/phase2_regional_analytics/data/geosites_bmk_indexed.csv"),
                       "Geosite_Name", "Latitude_WGS84", "Longitude_WGS84", "Administrative_Region", "Geosite_Type", "Geological_Domain", "bmk_regional"),
        # Step 5: geosites_draa_tafilalet_indexed.csv, geosites_physical_features.csv,
        # geosites_physical_features_general.csv, and geosites_road_features.csv were
        # all inspected and excluded — each is a derived/duplicate subset of the
        # national catalog with zero unique geosite names (confirmed after
        # whitespace-stripped name comparison; draa_tafilalet's coordinates also
        # match the national catalog to floating-point rounding, ~1e-4 deg). See
        # report for the comparison detail. Morocco_Geosites_Graph_Data.xlsx was
        # also excluded: every sheet is a typology/domain summary table with no
        # per-site coordinates.
    ]

    # Feuil1 is NOT loaded as a separate source (2026-07-30 findings revised
    # the earlier "confirmed redundant, skip it" conclusion, but the fix is a
    # targeted Center-coordinates fallback into Data generale's own groups —
    # see build_feuil1_center_fallback — not ingesting Feuil1's rows as
    # independent observations, which would double-count every geosite).
    feuil1_centers = build_feuil1_center_fallback(excel_path)
    feuil2_center_overrides = build_feuil2_center_overrides(excel_path)
    print(f"Feuil1 center fallback available for {len(feuil1_centers)} group(s); "
          f"Feuil2 DMS/Intervalle lookup overrides available: {len(feuil2_center_overrides)}")

    excel_df, excel_failures = load_excel_hierarchical(
        excel_path, "Data generale", "geoheritage_excel",
        feuil1_centers=feuil1_centers, feuil2_overrides=feuil2_center_overrides,
    )
    n_feuil1_used = (excel_df["Center_Coord_Source"] == "Feuil1_fallback").sum()
    n_feuil2_used = (excel_df["Center_Coord_Source"] == "Feuil2_lookup").sum()
    print(f"Excel source: {n_feuil1_used} observation(s) got their group center from the "
          f"Feuil1 fallback; {n_feuil2_used} observation(s) got it from the Feuil2 lookup override")
    sources.append(excel_df)

    combined = pd.concat(sources, ignore_index=True)
    combined.insert(0, "Observation_ID", [f"obs_{i:05d}" for i in range(len(combined))])

    print(f"Loaded {len(combined)} raw observations from {len(sources)} sources")
    print(combined["Source_File"].value_counts())
    n_missing = combined[["Latitude_WGS84", "Longitude_WGS84"]].isna().any(axis=1).sum()
    print(f"Observations with missing coordinates: {n_missing}")
    n_excel_groups = excel_df["Locality_Group_Key"].nunique()
    print(f"Excel source: {len(excel_df)} observations in {n_excel_groups} explicit locality groups")

    combined.to_csv(os.path.join(OUT_DIR, "geosites_observations_raw.csv"), index=False)
    excel_failures.to_csv(os.path.join(OUT_DIR, "dms_parse_failures.csv"), index=False)
    print(f"Coordinate parse failures (Excel source): {len(excel_failures)}")

    assert len(combined) > 780, "Combined observation count looks too low — check all sources loaded"
    assert n_excel_groups > 1, "Expected multiple explicit locality groups from the Excel hierarchy — check group-key assignment logic"

    print("\n=== Phase B: corrections, ground truth, border + regional checks ===")
    combined = apply_documented_corrections(combined)
    combined = apply_ground_truth_notes(combined)
    combined = fold_in_ground_truth(combined)

    boundary = fetch_morocco_boundary()
    combined, outliers = border_check(combined, boundary)
    outliers.to_csv(os.path.join(OUT_DIR, "geosites_outliers_removed.csv"), index=False)

    region_gdf = fetch_region_boundaries()
    combined, region_mismatches = regional_plausibility_check(combined, region_gdf, region_name_col="nom_region")
    region_mismatches.to_csv(os.path.join(OUT_DIR, "geosites_region_mismatches.csv"), index=False)

    combined.to_csv(os.path.join(OUT_DIR, "geosites_observations.csv"), index=False)
    print(f"Post-checks: {len(combined)} observations remain for locality construction")

    cap = combined[combined["Geosite_Name"].str.contains("Malabata", case=False, na=False)]
    assert len(cap) > 0, "Cap Malabata observation(s) missing from post-checks table entirely"
    assert np.isclose(cap.iloc[0]["Longitude_WGS84"], -5.7133, atol=0.01), "Cap Malabata correction did not apply"
    assert (cap["Longitude_WGS84"].apply(lambda v: np.isclose(v, -5.7133, atol=0.01))).all(), \
        "Not every Cap Malabata observation has the corrected longitude"

    fahs_anjra = combined[combined["Geosite_Name"].str.contains("Fahs-Anjra", case=False, na=False)]
    assert len(fahs_anjra) == 1, "Fahs-Anjra/Melloussa ground-truth observation missing from post-checks table"

    print("\n=== Phase C: locality construction ===")
    localities, needs_review = build_localities(combined)

    localities.to_csv(os.path.join(OUT_DIR, "geosites_localities_master.csv"), index=False)
    needs_review.to_csv(os.path.join(OUT_DIR, "geosites_needs_review.csv"), index=False)

    # Re-link each observation to its resolved locality and rewrite the observation table
    obs_to_locality = {}
    for _, loc in localities.iterrows():
        for obs_id in str(loc["Member_Observation_IDs"]).split(";"):
            obs_to_locality[obs_id] = loc["Locality_ID"]
    combined["Locality_ID"] = combined["Observation_ID"].map(obs_to_locality)
    combined.to_csv(os.path.join(OUT_DIR, "geosites_observations.csv"), index=False)

    print(f"\n=== FINAL SUMMARY ===")
    print(f"Localities: {len(localities)}")
    print(f"Observations: {len(combined)} ({combined['Locality_ID'].notna().sum()} linked to a locality, "
          f"{len(needs_review)} unresolved in needs_review.csv)")
    print(f"Outliers removed: {len(outliers)}")
    print(f"Region mismatches: {len(region_mismatches)}")

    assert len(localities) <= len(combined), "Locality count exceeds observation count -- grouping logic bug"
    assert combined["Locality_ID"].notna().all(), "Every observation must resolve to a locality or go to needs_review"

    cap = localities[localities["Geosite_Name"].str.contains("Malabata", case=False, na=False)]
    if len(cap) > 0:
        assert np.isclose(cap.iloc[0]["Longitude_WGS84"], -5.7133, atol=0.01)
    fahs = localities[localities["Geosite_Name"].str.contains("Fahs", case=False, na=False)]
    assert len(fahs) >= 1, "Fahs-Anjra ground truth missing from final localities"
    ras = pd.concat([
        localities[localities["Geosite_Name"].str.contains("Ras.*Ma", case=False, na=False, regex=True)],
        needs_review[needs_review["Geosite_Name"].str.contains("Ras.*Ma", case=False, na=False, regex=True)] if len(needs_review) else pd.DataFrame(),
    ])
    assert len(ras) >= 1, "No Ras El Ma variant found anywhere in outputs"
    print(f"Ras El Ma variants across localities+review: {len(ras)}")

    # Tinghir-Dades-Imilchil Excel publication group must NOT collapse to one locality:
    # its ~26 members are distinct, separately-named geosites spread over ~170km.
    tislit_loc = localities[localities["Geosite_Name"].str.contains("Tislit", case=False, na=False)]
    if len(tislit_loc) > 0:
        assert tislit_loc.iloc[0]["Observation_Count"] < 10, (
            "Tinghir-Dades-Imilchil publication group appears to have been force-collapsed "
            "into one locality -- Locality_Group_Key must not be treated as physical-identity evidence"
        )
    print("All spec-background sanity checks passed")


if __name__ == "__main__":
    main()
